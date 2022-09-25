import sqlite3
from xlsxwriter.workbook import Workbook
import openpyxl


def export_file(database_name, output_file_name):
    workbook = Workbook(output_file_name)
    worksheet = workbook.add_worksheet()

    connection = sqlite3.connect(database_name)
    c = connection.cursor()
    mysel = c.execute('''SELECT u.Id, u.Second_name, u.First_name, u.Patronymic, r.Region_name, c.City_name, u.Phone, u.Email 
                         FROM users u
                         INNER JOIN regions r ON r.Id = u.Region_id 
                         INNER JOIN cities c ON c.Region_id = u.Region_id AND c.Id = u.City_id''')
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            worksheet.write(i, j, row[j])
    workbook.close()
    # сохраняем изменения
    connection.commit()
    connection.close()


def input_file(database_name, input_file_name):
    connection = sqlite3.connect(database_name)
    cursor = connection.cursor()
    file_to_read = openpyxl.load_workbook(input_file_name, data_only=True)
    sheet = file_to_read['Sheet1']
    cursor.execute('SELECT * FROM users')
    users = []
    for row in cursor:
        users.append(list(row))
    dict_region = {}
    cursor.execute('SELECT Id, Region_name FROM regions')
    for row in cursor:
        dict_region[row[1]] = row[0]

    dict_city = {}
    cursor.execute('SELECT Id, City_name FROM cities')
    for row in cursor:
        dict_city[row[1]] = row[0]
    for row in range(1, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 8 ( 9 не включая)
        for col in range(1, 9):
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)
            number_of_elements = len(data)
            if number_of_elements == 8:
                data[4] = dict_region[data[4]]
                data[5] = dict_city[data[5]]
                if data not in users:
                    Update_Table = """INSERT OR REPLACE INTO users VALUES(?, ?, ?, ?, ?, ?, ?, ?);"""
                    cursor.execute(Update_Table,
                                   (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7]))

    # сохраняем изменения
    connection.commit()
    connection.close()
